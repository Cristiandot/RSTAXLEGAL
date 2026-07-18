import openpyxl, json
wb = openpyxl.load_workbook(r"C:\Users\CristianLópezThienel\OneDrive - Rodríguez Samith Tax & Legal Limitada\RSTL - Clientes\C.20 - Mauricio Miranda\TEVICOMP\02-Contab\TEVICOMP_2026_v9.xlsx", data_only=True, read_only=True)
out = {}

# COMPRAS: detectar encabezados y sumar por periodo
ws = wb["COMPRAS"]
filas = ws.iter_rows(values_only=True)
enc = [str(c).strip() if c is not None else "" for c in next(filas)]
out["compras_encabezados"] = enc
idx = {n: i for i, n in enumerate(enc)}
def col(*names):
    for n in names:
        if n in idx: return idx[n]
    return None
iPer, iNeto, iIva, iTot, iTipo = col("Periodo"), col("Monto Neto"), col("IVA Recuperable","Monto IVA Recuperable"), col("Monto Total"), col("Tipo Doc")
acc = {}
for f in filas:
    if f is None or iPer is None or f[iPer] is None: continue
    p = str(f[iPer])[:6]
    if not p.startswith("2026"): continue
    a = acc.setdefault(p, {"n":0,"neto":0,"iva":0,"total":0})
    a["n"] += 1
    a["neto"] += int(f[iNeto] or 0)
    a["iva"] += int(f[iIva] or 0)
    a["total"] += int(f[iTot] or 0)
out["compras_por_periodo"] = acc

# HONORARIOS: sumar por periodo
ws = wb["HONORARIOS"]
filas = ws.iter_rows(values_only=True)
enc2 = [str(c).strip() if c is not None else "" for c in next(filas)]
out["honorarios_encabezados"] = enc2
idx2 = {n: i for i, n in enumerate(enc2)}
accH = {}
for f in filas:
    if f is None or f[0] is None: continue
    p = str(f[0])[:6]
    if not p.startswith("2026"): continue
    a = accH.setdefault(p, {"n":0,"brutos":0,"retencion":0,"liquido":0})
    a["n"] += 1
    a["brutos"] += int(f[idx2.get("Brutos",8)] or 0)
    a["retencion"] += int(f[idx2.get("Retención", idx2.get("Retencion",9))] or 0)
    a["liquido"] += int(f[idx2.get("Líquido Pagado", idx2.get("Liquido Pagado",10))] or 0)
out["honorarios_por_periodo"] = accH

with open(r"C:\Proyectos\RSTAXLEGAL\app\tmp-tevicomp\workbook-validacion.json","w",encoding="utf-8") as fh:
    json.dump(out, fh, ensure_ascii=False, indent=1)
print("OK")
