import openpyxl, json, glob, datetime
out = {"archivos": []}
for ruta in sorted(glob.glob(r"C:\Users\CristianLópezThienel\OneDrive - Rodríguez Samith Tax & Legal Limitada\RSTL - Clientes\C.20 - Mauricio Miranda\TEVICOMP\02-Contab\Honorarios\*.xlsx")):
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    enc = [str(c).strip() if c is not None else "" for c in filas[0]]
    docs = []
    for f in filas[1:]:
        if f is None or all(v is None or str(v).strip()=="" for v in f): continue
        d = {}
        for i, n in enumerate(enc):
            v = f[i] if i < len(f) else None
            if isinstance(v, datetime.datetime): v = v.strftime("%Y-%m-%d")
            d[n] = v
        docs.append(d)
    out["archivos"].append({"archivo": ruta.split("\\")[-1], "hojas": wb.sheetnames, "encabezados": enc, "n": len(docs), "docs": docs})
with open(r"C:\Proyectos\RSTAXLEGAL\app\tmp-tevicomp\honorarios.json","w",encoding="utf-8") as fh:
    json.dump(out, fh, ensure_ascii=False, indent=1)
print("archivos:", len(out["archivos"]), "| docs:", sum(a["n"] for a in out["archivos"]))
