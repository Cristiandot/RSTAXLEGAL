import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\CristianLópezThienel\OneDrive - Rodríguez Samith Tax & Legal Limitada\RSTL - Clientes\C.20 - Mauricio Miranda\TEVICOMP\02-Contab\TEVICOMP_2026_v9.xlsx", data_only=True, read_only=True)
print("HOJAS:", wb.sheetnames)
if "REPORTE" in wb.sheetnames:
    ws = wb["REPORTE"]
    for fila in ws.iter_rows(min_row=1, max_row=45, max_col=9, values_only=True):
        if any(v is not None and v != "" for v in fila):
            print(" | ".join("" if v is None else str(v) for v in fila))
